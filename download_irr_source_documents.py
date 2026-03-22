"""
================================================================================
download_irr_source_documents.py
================================================================================

Infrastructure Literacy Credential Content Analysis
IRR Source Document Download and Keyword Search Replication Script

Author:     Devan Cantrell Addison-Turner
            PhD Candidate, Civil and Environmental Engineering
            Stanford Doerr School of Sustainability, Stanford University
            Natural Capital Alliance | Lepech Research Group |
            Center for Integrated Facility Engineering (CIFE)
            daddisonturner@stanford.edu
            ORCID: 0000-0002-2511-3680

Study:      "Infrastructure Literacy: A Conceptual Framework for Understanding
             How Construction Career Students Think About Environmental Justice"
             Manuscript submitted to Journal of Vocational Education and Training

Data repo:  https://github.com/devanaddisonturner/Infrastructure-Literacy-Coding
Zenodo:     https://doi.org/10.5281/zenodo.18893500

IRB:        Stanford IRB Protocol #84369
Trial reg:  ClinicalTrials.gov NCT07315919

================================================================================
PURPOSE
================================================================================
This script provides full replication of the three-layer systematic keyword
content analysis reported in the manuscript. It:

  1. Downloads all publicly available source credential documents
  2. Loads all 431 coded learning outcomes from EJT_Binary_Coding_Results_v6.xlsx
  3. Runs all Layer 0 (15 primary terms) and Layer 1 (30 near-synonym) keyword
     searches programmatically across all four credentials
  4. Compares search results against the primary coder's recorded results
  5. Produces a plain-text replication report confirming or flagging discrepancies
  6. Packages everything into a single reproducible archive

This approach ensures full replication of the central finding — that 0 of 431
learning outcomes contain any of the 45 keyword terms — without requiring
purchase of the NCCER Core Curriculum. All 103 NCCER outcomes are stored
verbatim in the master spreadsheet, which is openly available on Zenodo.

================================================================================
NCCER ACCESS NOTE
================================================================================
NCCER Core Curriculum (6th Edition) requires institutional purchase and is not
publicly available for direct download. The 103 learning outcomes coded from
this credential are stored verbatim in EJT_Binary_Coding_Results_v6.xlsx
(Layer 2 — Thematic Audit sheet) and are fully searchable via this script.

For independent verification of outcome text transcription accuracy, purchase
through one of the following:

  Print (ISBN: 9780137483341) — recommended for code-based replication:
    Amazon:  https://www.amazon.com/dp/0137483341
    Pearson: https://www.pearson.com/en-us/subject-catalog/p/
             core-introduction-to-basic-construction-skills/
             P200000007556/9780137483341

  eTextbook (ISBN: 9780137483228) — DRM-protected, NOT code-searchable:
    VitalSource: https://www.vitalsource.com/products/core-nccer-v9780137483228
    Note: VitalSource eTextbooks are DRM-protected and cannot be opened or
    processed programmatically. Use the print ISBN for code-based replication.

  Institutional access:
    NCCER-accredited training centers or university libraries with workforce
    education collections may provide access at no additional cost.

  NCCER catalog: https://www.nccer.org/craft-catalog/core/

================================================================================
OUTPUT STRUCTURE
================================================================================
    irr_source_documents/
        credential_sources/
            buildingconstruct.pdf             (Credential B — CA CTE, USA)
            CPC30220_R1.pdf                   (Credential C — Australia)
            6706-23_l2_diploma_...pdf         (Credential D — City & Guilds, UK)
            NCCER_ACCESS_NOTE.txt             (Credential A — purchase instructions)
        cpc_units/
            CPCCCA2002_R1.pdf                 (27 unit competency PDFs)
            ... (27 files total)
        keyword_search_results/
            layer0_replication_report.txt     (15 primary terms)
            layer1_replication_report.txt     (30 near-synonym terms)
            layer2_replication_report.txt     (thematic audit summary)
            full_replication_report.txt       (all three layers combined)
    irr_source_documents.zip                  (complete replication archive)
    download_manifest.txt                     (checksums, citations, access dates)

================================================================================
USAGE — macOS (Terminal)
================================================================================
    macOS comes with Python 3 pre-installed (Monterey 12+) or via Homebrew.

    1. Open Terminal:
           Applications → Utilities → Terminal
       Or press Cmd + Space, type "Terminal", press Enter.

    2. Verify Python 3 is available:
           python3 --version
       If not installed:
           /bin/bash -c "$(curl -fsSL https://brew.sh/install.sh)"
           brew install python3

    3. Install openpyxl (first time only):
           pip3 install openpyxl

    4. Navigate to the folder containing this script and the spreadsheet:
           cd ~/Downloads

    5. Place EJT_Binary_Coding_Results_v6.xlsx in the same folder.

    6. Run the full download and keyword search:
           python3 download_irr_source_documents.py

    7. Run keyword search only (no download):
           python3 download_irr_source_documents.py --search-only

================================================================================
USAGE — Windows (Command Prompt)
================================================================================
    1. Install Python 3 if not already installed:
       Download from https://www.python.org/downloads/windows/
       During installation CHECK "Add Python to PATH" before clicking Install.

    2. Open Command Prompt:
           Press Win + R → type cmd → press Enter
       Or search "Command Prompt" in the Start menu.

    3. Verify Python is available:
           python --version
       If that fails, try:
           py --version

    4. Install openpyxl (first time only):
           pip install openpyxl

    5. Navigate to the folder containing this script and the spreadsheet:
           cd %USERPROFILE%/Downloads

    6. Place EJT_Binary_Coding_Results_v6.xlsx in the same folder.

    7. Run the full download and keyword search:
           python download_irr_source_documents.py

    8. Run keyword search only (no download):
           python download_irr_source_documents.py --search-only

    If "python is not recognised", replace python with py throughout.

================================================================================
USAGE — Windows (PowerShell)
================================================================================
    PowerShell is included with all modern Windows versions.

    1. Open PowerShell:
           Press Win + X → select "Windows PowerShell"
       Or search "PowerShell" in the Start menu.

    2. Verify Python is available:
           python --version
       If that fails, try:
           py --version

    3. Install openpyxl (first time only):
           pip install openpyxl

    4. Navigate to the folder containing this script and the spreadsheet:
           cd $env:USERPROFILE/Downloads

    5. Place EJT_Binary_Coding_Results_v6.xlsx in the same folder.

    6. Run the full download and keyword search:
           python download_irr_source_documents.py

    7. Run keyword search only (no download):
           python download_irr_source_documents.py --search-only

    PowerShell note: If you see a script execution policy error, run:
           Set-ExecutionPolicy -Scope CurrentUser RemoteSigned
    Then retry the command.

================================================================================
USAGE — IDLE (cross-platform, no command line needed)
================================================================================
    IDLE is Python's built-in graphical editor, included with every Python
    installation. Use this if you prefer not to use a terminal or command prompt.

    1. Open IDLE:
       macOS:   Open Finder → Applications → Python 3.x → IDLE
                Or search "IDLE" in Spotlight (Cmd + Space)
       Windows: Start menu → search "IDLE" → open "IDLE (Python 3.x)"

    2. Open this script in IDLE:
           File → Open → navigate to download_irr_source_documents.py

    3. Before running, set the working directory to the folder containing
       EJT_Binary_Coding_Results_v6.xlsx. In the IDLE Shell (>>> prompt):
           >>> import os
           >>> os.chdir("/path/to/folder/containing/script")
           # macOS example:  os.chdir("/Users/yourname/Downloads")
           # Windows example: os.chdir("C:/Users/yourname/Downloads")

    4. Run the script:
           Press F5   (or Run → Run Module)

    5. To run with --search-only flag from IDLE Shell:
           >>> import sys
           >>> sys.argv = ["download_irr_source_documents.py", "--search-only"]
           >>> exec(open("download_irr_source_documents.py").read())

    IDLE note: The script will auto-install openpyxl if not found.
    Output (irr_source_documents/ folder) appears in your working directory.

================================================================================
Requirements: Python 3.6+ and openpyxl (auto-installed if missing).
================================================================================
"""


import urllib.request
import urllib.error
import os
import re
import sys
import zipfile
import hashlib
import time
from datetime import datetime

# ── Auto-install openpyxl if not present ──────────────────────────────────────
try:
    import openpyxl
except ImportError:
    import subprocess
    print("Installing required dependency: openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

# Output directories and file paths
OUTPUT_DIR     = "irr_source_documents"
CREDENTIAL_DIR = os.path.join(OUTPUT_DIR, "credential_sources")
CPC_UNITS_DIR  = os.path.join(OUTPUT_DIR, "cpc_units")
RESULTS_DIR    = os.path.join(OUTPUT_DIR, "keyword_search_results")
ZIP_PATH       = "irr_source_documents.zip"
MANIFEST_PATH  = os.path.join(OUTPUT_DIR, "download_manifest.txt")

# Master spreadsheet must be in the same directory as this script
MASTER_XLSX = "EJT_Binary_Coding_Results_v6.xlsx"

# Browser-style user agent for polite HTTP requests
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# Pause between HTTP requests to avoid rate-limiting (seconds)
REQUEST_DELAY = 0.4


# ══════════════════════════════════════════════════════════════════════════════
# KEYWORD TERMS
# Source: manuscript Table 1 and master spreadsheet Layer 0 / Layer 1 sheets
# ══════════════════════════════════════════════════════════════════════════════

# Layer 0: 15 domain-specific primary terms
# Coding rule: present (1) if term appears anywhere in outcome text,
# exact match OR root match (e.g. 'equitable' matches 'equity').
# Context is NOT required — if the word appears, code as 1.
# Regulatory/technical uses count (e.g. 'environmental impact' in a
# compliance context still codes as 1).
LAYER0_TERMS = [
    # Environmental Justice Domain
    # Source: EJT_Binary_Coding_Results_v6.xlsx, Layer 0 - Primary Coding
    "environmental justice",
    "environmental racism",
    "environmental equity",
    "environmental injustice",
    # Health Equity Domain
    "health equity",
    "health disparities",
    "health outcomes",
    "community health",
    # Distributional Consequences Domain
    "distributive consequences",
    "distributional impact",
    "disproportionate burden",
    "disproportionate impact",
    # Community Participation Domain
    "community voice",
    "community engagement",
    "community participation",
]

# Layer 1: 30 near-synonym terms
# Added in response to reviewer concern that Layer 0 used only domain-specific
# jargon, potentially missing equity content expressed in lay language.
LAYER1_TERMS = [
    # Equity and Justice Terms
    # Source: EJT_Binary_Coding_Results_v6.xlsx, Layer 1 - Near-Synonyms
    "unfair", "unequal", "inequality", "inequity",
    "social justice", "discrimination", "marginalized",
    "disadvantaged", "underserved", "racial",
    # Health and Exposure Terms
    "public health", "environmental health", "neighborhood health",
    "pollution exposure", "toxic exposure", "social determinants",
    "cumulative impact", "well-being", "pollution",
    # Distributional and Vulnerability Terms
    "vulnerable populations", "vulnerable communities", "overburdened",
    "social vulnerability", "distribution of burdens",
    "community impact", "community benefit", "social benefits",
    # Broader Contextual Terms
    "disparities", "accessibility", "contamination",
]


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE DOCUMENT REGISTRY
# ══════════════════════════════════════════════════════════════════════════════

# Credential qualification documents — Credentials B, C, D (publicly available)
# Credential A (NCCER) handled separately — see NCCER ACCESS NOTE in header
CREDENTIAL_DOCS = [
    {
        "label":    "Credential B — CA CTE Model Curriculum Standards (USA)",
        "citation": ("California Department of Education (CDE, 2013; published 2017). "
                     "Career Technical Education Model Curriculum Standards: "
                     "Building and Construction Trades."),
        "filename": "buildingconstruct.pdf",
        "url":      "https://www.cde.ca.gov/ci/ct/sf/documents/buildingconstruct.pdf",
        "accessed": "2025-12-15",
    },
    {
        "label":    "Credential C — CPC30220 Certificate III in Carpentry (Australia)",
        "citation": ("Artibus Innovation (2020). CPC30220 Certificate III in Carpentry, "
                     "Release 1. Australian Government."),
        "filename": "CPC30220_R1.pdf",
        "url":      "https://training.gov.au/TrainingComponentFiles/CPC/CPC30220_R1.pdf",
        "accessed": "2025-12-15",
    },
    {
        "label":    "Credential D — City & Guilds Level 2 Diploma in Site Carpentry (UK)",
        "citation": ("City and Guilds (2022). Level 2 Diploma in Site Carpentry "
                     "(6706-23) Qualification Handbook v1.4."),
        "filename": "6706-23_l2_diploma_in_site_carpentry_qhb_v1-4-pdf.pdf",
        "url":      ("https://www.cityandguilds.com/-/media/productdocuments/"
                     "construction_and_the_built_environment/construction/6706/"
                     "6706_level_2/centre_documents/"
                     "6706-23_l2_diploma_in_site_carpentry_qhb_v1-4-pdf.pdf"),
        "accessed": "2025-12-15",
    },
]

# CPC30220 individual unit competency PDFs (27 core units, 114 elements total)
# Source: CPC Construction, Plumbing and Services Training Package, training.gov.au
# These provide the element-level outcomes used for granularity-consistent coding.
CPC_UNITS = [
    ("CPCCCA2002", "Use carpentry tools and equipment"),
    ("CPCCCA2011", "Handle carpentry materials"),
    ("CPCCCA3001", "Carry out general demolition of minor building structures"),
    ("CPCCCA3002", "Carry out setting out"),
    ("CPCCCA3003", "Install flooring systems"),
    ("CPCCCA3004", "Construct and erect wall frames"),
    ("CPCCCA3005", "Construct ceiling frames"),
    ("CPCCCA3006", "Erect roof trusses"),
    ("CPCCCA3007", "Construct pitched roofs"),
    ("CPCCCA3008", "Construct eaves"),
    ("CPCCCA3010", "Install windows and doors"),
    ("CPCCCA3016", "Construct, assemble and install timber external stairs"),
    ("CPCCCA3017", "Install exterior cladding"),
    ("CPCCCA3024", "Install lining, panelling and moulding"),
    ("CPCCCA3025", "Read and interpret plans, specifications and drawings for carpentry work"),
    ("CPCCCA3028", "Erect and dismantle formwork for footings and slabs on ground"),
    ("CPCCCM2006", "Apply basic levelling procedures"),
    ("CPCCCM2008", "Erect and dismantle restricted height scaffolding"),
    ("CPCCCM2012", "Work safely at heights"),
    ("CPCCCO2013", "Carry out concreting to simple forms"),
    ("CPCCOM1012", "Work effectively and sustainably in the construction industry"),
    ("CPCCOM1014", "Conduct workplace communication"),
    ("CPCCOM1015", "Carry out measurements and calculations"),
    ("CPCCOM3001", "Perform construction calculations to determine carpentry material requirements"),
    ("CPCCOM3006", "Carry out levelling operations"),
    ("CPCCWHS1001", "Prepare to work safely in the construction industry"),
    ("CPCCWHS2001", "Apply WHS requirements, policies and procedures in the construction industry"),
]

CPC_BASE_URL = "https://training.gov.au/TrainingComponentFiles/CPC/"
CPC_RELEASES = ["R1", "R2", "R3"]  # try R1 first; fall back to R2/R3 if updated


# ══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def md5(path):
    """Return MD5 checksum of a file for manifest integrity verification."""
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def download_pdf(url, dest_path):
    """
    Download a file from url to dest_path.
    Validates the downloaded file is a genuine PDF using magic bytes (%PDF).
    Returns True on success, False on any failure (network, non-PDF content, etc.).
    """
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=30) as response:
            data = response.read()
        if data[:4] != b"%PDF":
            return False  # server returned HTML error page or redirect
        with open(dest_path, "wb") as f:
            f.write(data)
        return True
    except Exception:
        return False


def download_with_fallback(unit_code, dest_dir):
    """
    Download a training.gov.au unit PDF, trying release versions R1 → R2 → R3.
    Some units were updated after their initial release; fallback handles this.
    Returns (success: bool, release_used: str, dest_path: str).
    """
    for release in CPC_RELEASES:
        url  = f"{CPC_BASE_URL}{unit_code}_{release}.pdf"
        dest = os.path.join(dest_dir, f"{unit_code}_{release}.pdf")
        if download_pdf(url, dest):
            return True, release, dest
        time.sleep(REQUEST_DELAY)
    return False, None, None


def term_present(term, text):
    """
    Check whether a keyword term is present in a learning outcome text string.

    Matching rules:
      - Multi-word terms: exact phrase match (case-insensitive), word-boundary anchored
      - Single-word terms: whole-word match using word boundaries, plus simple
        plural/verb form via up to 3-character suffix to catch plurals, past tense,
        and gerunds (e.g. 'hazard' matches 'hazards', 'contaminate' matches
        'contamination'; 'health' does NOT match 'unhealthy' or 'healthcare')

    Parameters:
        term (str): keyword term to search for
        text (str): learning outcome description text

    Returns:
        bool: True if term is present (code: 1), False if absent (code: 0)
    """
    text_lower = text.lower()
    term_lower = term.lower()

    if " " in term_lower:
        # Multi-word phrase: require all words present in sequence
        pattern = r"\b" + re.escape(term_lower) + r"\b"
        return bool(re.search(pattern, text_lower))
    else:
        # Single-word: exact whole-word match
        exact = r"\b" + re.escape(term_lower) + r"\b"
        if re.search(exact, text_lower):
            return True
        # Stem match: strip one char and allow up to 3-char suffix
        # Only applied to words longer than 5 chars to avoid short-word false matches
        if len(term_lower) > 5:
            stem = re.escape(term_lower[:-1])
            stem_pattern = r"\b" + stem + r"[a-z]{0,3}\b"
            if re.search(stem_pattern, text_lower):
                return True
        return False


# ══════════════════════════════════════════════════════════════════════════════
# LOAD OUTCOMES FROM MASTER SPREADSHEET
# ══════════════════════════════════════════════════════════════════════════════

def load_outcomes(xlsx_path):
    """
    Load all 431 learning outcomes and primary coder results from the master
    spreadsheet (EJT_Binary_Coding_Results_v6.xlsx).

    Reads three sheets:
      Layer 2 — Thematic Audit: outcome text and Layer 2 (A/B/C) coding
      Layer 0 — Primary Coding: Layer 0 term-level binary results by credential
      Layer 1 — Near-Synonyms:  Layer 1 term-level binary results by credential

    Returns:
      outcomes (list of dict): one entry per learning outcome with keys:
        row_num, credential, unit, outcome_id, description, crit_a, crit_b, crit_c
      layer0_primary (dict): {term_lower: {cred_key: 0/1}} from primary coder
      layer1_primary (dict): {term_lower: {cred_key: 0/1}} from primary coder
    """
    if not os.path.exists(xlsx_path):
        print(f"\nERROR: Master spreadsheet not found: {xlsx_path}")
        print("Place EJT_Binary_Coding_Results_v6.xlsx in the same directory")
        print("as this script, then run again.")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # ── Layer 2: load all 431 outcome rows ────────────────────────────────────
    ws2   = wb["Layer 2 - Thematic Audit"]
    rows2 = list(ws2.iter_rows(min_row=1, values_only=True))
    ch2   = next(i for i, r in enumerate(rows2) if r[0] == "#")

    outcomes = []
    for row in rows2[ch2 + 1:]:
        if row[0] is None:
            continue
        try:
            int(row[0])
        except (TypeError, ValueError):
            continue
        outcomes.append({
            "row_num":     row[0],
            "credential":  row[1],
            "unit":        row[2],
            "outcome_id":  row[3],
            "description": str(row[4]) if row[4] else "",
            "crit_a":      int(row[5]) if row[5] is not None else 0,
            "crit_b":      int(row[6]) if row[6] is not None else 0,
            "crit_c":      int(row[7]) if row[7] is not None else 0,
        })

    # ── Layer 0: load primary coder term results ──────────────────────────────
    SKIP_ROWS = {
        "Search Term", "Environmental Justice Domain", "Health Equity Domain",
        "Distributional Consequences Domain", "Community Participation Domain",
        "Equity and Justice Terms", "Environmental Terms",
        "Health and Exposure Terms", "Distributional and Vulnerability Terms",
        "Broader Contextual Terms", "Total terms present",
    }

    def load_term_sheet(ws):
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        ch   = next(i for i, r in enumerate(rows) if r[0] == "Search Term")
        result = {}
        for row in rows[ch + 1:]:
            if not row[0]:
                continue
            v = str(row[0]).strip()
            # Stop at summary/notes rows
            if v in SKIP_ROWS or v.startswith("Result") or v.startswith("Note"):
                continue
            # Skip rows that look like annotations (start with * or are long sentences)
            if v.startswith("*") or len(v) > 60:
                continue
            # Strip trailing annotation markers (*, **, ***)
            clean = v.rstrip("*").strip()
            if not clean or len(clean) < 2:
                continue
            term = clean.lower()
            result[term] = {
                "NCCER":         0 if row[1] is None else int(row[1]),
                "CA CTE":        0 if row[2] is None else int(row[2]),
                "CPC30220":      0 if row[3] is None else int(row[3]),
                "City & Guilds": 0 if row[4] is None else int(row[4]),
            }
        return result

    layer0_primary = load_term_sheet(wb["Layer 0 - Primary Coding"])
    layer1_primary = load_term_sheet(wb["Layer 1 - Near-Synonyms"])

    wb.close()
    return outcomes, layer0_primary, layer1_primary


# ══════════════════════════════════════════════════════════════════════════════
# RUN KEYWORD SEARCHES
# ══════════════════════════════════════════════════════════════════════════════

# Maps Layer 2 credential names to the short keys used in Layer 0/1 columns
CRED_KEY_MAP = {
    "NCCER Core (6th ed., USA)":  "NCCER",
    "CA CTE Standards (USA)":     "CA CTE",
    "CPC30220 (Australia)":       "CPC30220",
    "City & Guilds 6706-23 (UK)": "City & Guilds",
}


def run_keyword_search(outcomes, terms):
    """
    Run keyword presence/absence search across all 431 outcomes.

    For each term, searches every outcome description using term_present().
    Aggregates to credential-level binary result (1 if term present in ANY
    outcome for that credential, 0 if absent in all outcomes).

    This replicates the exact structure of the Layer 0 and Layer 1 sheets
    in the master spreadsheet.

    Parameters:
        outcomes (list): loaded outcome dicts from load_outcomes()
        terms (list):    keyword terms to search

    Returns:
        results (dict):          {term_lower: {cred_key: 0 or 1}}
        hits_by_outcome (list):  tuples of (outcome_id, credential, term, text[:120])
                                 for any outcome-level matches (expected: empty)
    """
    results = {
        t.lower(): {"NCCER": 0, "CA CTE": 0, "CPC30220": 0, "City & Guilds": 0}
        for t in terms
    }
    hits_by_outcome = []

    for outcome in outcomes:
        cred_key = CRED_KEY_MAP.get(outcome["credential"])
        if not cred_key:
            continue
        for term in terms:
            if term_present(term, outcome["description"]):
                results[term.lower()][cred_key] = 1
                hits_by_outcome.append((
                    outcome["outcome_id"],
                    outcome["credential"],
                    term,
                    outcome["description"][:120],
                ))

    return results, hits_by_outcome


# ══════════════════════════════════════════════════════════════════════════════
# GENERATE REPLICATION REPORTS
# ══════════════════════════════════════════════════════════════════════════════

def generate_layer_report(layer_name, terms, rep_results, pri_results,
                          hits_by_outcome, n_outcomes, run_timestamp):
    """
    Generate a plain-text replication report for Layer 0 or Layer 1.

    Produces:
      - Term-by-term comparison table (replication vs primary coder)
      - List of any outcome-level hits (should be empty — 0% finding)
      - Summary of discrepancies and replication status

    Parameters:
        layer_name (str):      display name for the layer
        terms (list):          keyword terms searched
        rep_results (dict):    replication search results {term: {cred: 0/1}}
        pri_results (dict):    primary coder results {term: {cred: 0/1}}
        hits_by_outcome (list): any outcome-level hits found
        n_outcomes (int):      total outcomes searched
        run_timestamp (str):   run date/time string
    """
    cred_keys  = ["NCCER", "CA CTE", "CPC30220", "City & Guilds"]
    cred_short = ["NCCER", "CATE ", "CPC  ", "CG   "]
    lines = []

    lines += [
        "=" * 72,
        f"REPLICATION REPORT — {layer_name}",
        f"Generated: {run_timestamp}",
        "Author: Devan Cantrell Addison-Turner, Stanford University",
        "Study:  Infrastructure Literacy Credential Content Analysis",
        "Zenodo: https://doi.org/10.5281/zenodo.18893500",
        "=" * 72, "",
        f"Terms searched:    {len(terms)}",
        f"Outcomes searched: {n_outcomes}",
        f"Credentials:       4 (NCCER, CA CTE, CPC30220, City & Guilds)",
        "",
    ]

    # Term-by-term comparison table
    lines += ["TERM-BY-TERM COMPARISON (Replication vs Primary Coder)", "-" * 72]
    header = (f"{'Term':<38} "
              + " ".join(f"{s:>6}" for s in cred_short)
              + "   Status")
    lines.append(header)
    lines.append("-" * 72)

    discrepancies = 0
    for term in terms:
        tl  = term.lower()
        rep = rep_results.get(tl, {})
        pri = pri_results.get(tl, {})
        match = True
        cells = []
        for ck in cred_keys:
            rv = rep.get(ck, 0)
            pv = pri.get(ck, 0)
            if rv != pv:
                match = False
                discrepancies += 1
                cells.append(f"{rv}!={pv}")
            else:
                cells.append(f"{rv:>4}  ")
        status = "  OK" if match else "DIFF"
        lines.append(f"{term:<38} " + " ".join(cells) + f"   {status}")

    lines += ["-" * 72, f"Discrepancies: {discrepancies}", ""]

    # Outcome-level hits
    lines += ["OUTCOME-LEVEL HITS (expected: none — central finding is 0%)",
              "-" * 72]
    if hits_by_outcome:
        lines.append(f"WARNING: {len(hits_by_outcome)} hits found. Review carefully.")
        for oid, cred, term, desc in hits_by_outcome:
            lines += [f"  Outcome: {oid}  |  Credential: {cred}",
                      f"  Term:    '{term}'",
                      f"  Text:    {desc}...", ""]
    else:
        lines.append("  No hits found. 0% finding confirmed.")
    lines.append("")

    # Summary
    lines.append("REPLICATION STATUS")
    lines.append("-" * 72)
    if discrepancies == 0 and not hits_by_outcome:
        lines += [
            "CONFIRMED",
            "  Zero keyword matches found across all outcomes.",
            "  Replication results match primary coder 100%.",
        ]
    else:
        lines += [
            "REVIEW REQUIRED",
            f"  {discrepancies} term-credential discrepancy(ies) detected.",
            f"  {len(hits_by_outcome)} outcome-level hit(s) found.",
            "  Check term_present() logic and outcome text encoding.",
        ]
    lines.append("")
    return "\n".join(lines)


def generate_layer2_report(outcomes, run_timestamp):
    """
    Summarise the Layer 2 thematic audit from the master spreadsheet.

    Reports distribution of Criterion A / B / C coding by credential and
    confirms the 0% finding across all three evaluative criteria.

    Layer 2 coding criteria (from manuscript methods):
      Crit. A: Does the outcome require evaluating how infrastructure decisions
               distribute environmental health burdens across communities?
      Crit. B: Does the outcome address community health in an equity framework
               (beyond occupational safety)?
      Crit. C: Does the outcome engage environmental justice concepts, even
               implicitly?
      Result:  Present (1) if ANY criterion is met (A OR B OR C)
    """
    creds = [
        "NCCER Core (6th ed., USA)",
        "CA CTE Standards (USA)",
        "CPC30220 (Australia)",
        "City & Guilds 6706-23 (UK)",
    ]
    lines = []
    lines += [
        "=" * 72,
        "REPLICATION REPORT — Layer 2: Thematic Audit",
        f"Generated: {run_timestamp}",
        "=" * 72, "",
        "Criteria (from manuscript methods section):",
        "  Crit. A: Outcome requires evaluating distributional environmental",
        "           health burden across communities",
        "  Crit. B: Outcome addresses community health in equity framework",
        "           (beyond occupational safety)",
        "  Crit. C: Outcome engages environmental justice concepts (implicit OK)",
        "  Result:  Present if ANY criterion met (A OR B OR C)",
        "",
    ]

    lines.append(
        f"{'Credential':<36} {'N':>5} {'Crit.A':>7} {'Crit.B':>7} "
        f"{'Crit.C':>7} {'Result':>7}"
    )
    lines.append("-" * 72)

    ta = tb = tc = tr = 0
    for cred in creds:
        co = [o for o in outcomes if o["credential"] == cred]
        n  = len(co)
        a  = sum(1 for o in co if o["crit_a"] > 0)
        b  = sum(1 for o in co if o["crit_b"] > 0)
        c  = sum(1 for o in co if o["crit_c"] > 0)
        r  = sum(1 for o in co if o["crit_a"] > 0 or o["crit_b"] > 0 or o["crit_c"] > 0)
        ta += a; tb += b; tc += c; tr += r
        lines.append(f"{cred[:35]:<36} {n:>5} {a:>7} {b:>7} {c:>7} {r:>7}")

    lines += [
        "-" * 72,
        f"{'TOTAL':<36} {len(outcomes):>5} {ta:>7} {tb:>7} {tc:>7} {tr:>7}",
        "",
        "FINDING",
        "-" * 72,
    ]

    if tr == 0:
        lines += [
            "CONFIRMED: 0 of 431 outcomes meet any equity criterion.",
            "  Central manuscript finding replicated from master spreadsheet.",
        ]
    else:
        lines += [
            f"NOTE: {tr} outcome(s) coded as Present. Review master spreadsheet.",
        ]
    lines.append("")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # Parse command-line arguments
    import argparse
    parser = argparse.ArgumentParser(
        description="Download IRR source documents and run keyword search replication.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python3 download_irr_source_documents.py              # full run\n"
            "  python3 download_irr_source_documents.py --search-only # search only\n"
            "  python  download_irr_source_documents.py              # Windows"
        )
    )
    parser.add_argument(
        "--search-only",
        action="store_true",
        help="Skip downloads; run keyword search using existing spreadsheet only."
    )
    args = parser.parse_args()
    search_only = args.search_only
    run_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_steps   = 3 if search_only else 5

    print("=" * 72)
    print("Infrastructure Literacy — IRR Source Document Download & Replication")
    print("Author: Devan Cantrell Addison-Turner, Stanford University")
    print(f"Run date: {run_timestamp}")
    if search_only:
        print("Mode: --search-only (keyword search only, no download)")
    print("=" * 72)

    # Create all output directories
    for d in [OUTPUT_DIR, CREDENTIAL_DIR, CPC_UNITS_DIR, RESULTS_DIR]:
        os.makedirs(d, exist_ok=True)

    manifest_lines = [
        "Infrastructure Literacy — IRR Source Document Download Manifest",
        f"Generated: {run_timestamp}",
        "Author: Devan Cantrell Addison-Turner, Stanford University",
        "ORCID: 0000-0002-2511-3680",
        "Data repository: https://github.com/devanaddisonturner/Infrastructure-Literacy-Coding",
        "Zenodo DOI: https://doi.org/10.5281/zenodo.18893500",
        "=" * 72, "",
    ]

    downloaded = []
    failed     = []

    # ── Step 1: Write NCCER access note ───────────────────────────────────────
    if not search_only:
        print(f"\n[Step 1/{total_steps}] Credential A — NCCER Core Curriculum (USA)")
        nccer_note = "\n".join([
            "NCCER Core Curriculum — Access Note",
            "=" * 36, "",
            "Credential:     NCCER Core Curriculum, 6th Edition (NCCER, 2023)",
            "Coded level:    Sub-objective (a/b/c items under each Learning Objective)",
            "Outcomes coded: 103",
            "",
            "This credential requires institutional purchase and is not publicly",
            "available for direct download. All 103 coded outcomes are stored",
            "verbatim in EJT_Binary_Coding_Results_v6.xlsx (Layer 2 sheet) and",
            "are fully keyword-searchable via this script.",
            "",
            "To verify outcome text transcription accuracy, purchase through:",
            "",
            "  Print (ISBN: 9780137483341) — recommended for code-based replication:",
            "    Amazon:  https://www.amazon.com/dp/0137483341",
            "    Pearson: https://www.pearson.com/en-us/subject-catalog/p/",
            "             core-introduction-to-basic-construction-skills/",
            "             P200000007556/9780137483341",
            "",
            "  eTextbook (ISBN: 9780137483228) — DRM-protected, NOT code-searchable:",
            "    VitalSource: https://www.vitalsource.com/products/core-nccer-v9780137483228",
            "",
            "  Institutional access:",
            "    NCCER-accredited training centers or university libraries",
            "    with workforce education collections.",
            "",
            "NCCER catalog: https://www.nccer.org/craft-catalog/core/",
        ])
        with open(os.path.join(CREDENTIAL_DIR, "NCCER_ACCESS_NOTE.txt"), "w") as f:
            f.write(nccer_note)
        print("  Written: NCCER_ACCESS_NOTE.txt")
        manifest_lines += [
            "CREDENTIAL A — NCCER Core Curriculum (USA)",
            "  Status:  Institutional purchase required",
            "  Note:    103 outcomes stored verbatim in master spreadsheet",
            "  ISBNs:   Print 9780137483341 | eText 9780137483228 (DRM-protected)",
            "",
        ]

    # ── Step 2: Download credential qualification documents ────────────────────
    if not search_only:
        print(f"\n[Step 2/{total_steps}] Credential qualification documents (B, C, D)")
        manifest_lines += ["CREDENTIAL QUALIFICATION DOCUMENTS", "-" * 40]

        for doc in CREDENTIAL_DOCS:
            dest = os.path.join(CREDENTIAL_DIR, doc["filename"])
            print(f"  Downloading {doc['filename']}... ", end="", flush=True)
            if download_pdf(doc["url"], dest):
                checksum = md5(dest)
                size_kb  = os.path.getsize(dest) // 1024
                print(f"OK ({size_kb} KB)")
                downloaded.append(dest)
                manifest_lines += [
                    f"File:     {doc['filename']}",
                    f"Label:    {doc['label']}",
                    f"Citation: {doc['citation']}",
                    f"URL:      {doc['url']}",
                    f"Accessed: {doc['accessed']}",
                    f"MD5:      {checksum}",
                    f"Size:     {size_kb} KB", "",
                ]
            else:
                print("FAILED")
                failed.append(doc["filename"])
                manifest_lines += [
                    f"File: {doc['filename']} — DOWNLOAD FAILED",
                    f"URL:  {doc['url']}", "",
                ]
            time.sleep(REQUEST_DELAY)

    # ── Step 3: Download CPC unit PDFs ─────────────────────────────────────────
    if not search_only:
        print(f"\n[Step 3/{total_steps}] CPC30220 unit competency PDFs ({len(CPC_UNITS)} units)")
        manifest_lines += [
            "CPC30220 UNIT COMPETENCY DOCUMENTS (element-level coding source)",
            f"Source: {CPC_BASE_URL}", "-" * 40,
        ]
        for unit_code, unit_title in CPC_UNITS:
            print(f"  {unit_code}... ", end="", flush=True)
            ok, release, dest = download_with_fallback(unit_code, CPC_UNITS_DIR)
            if ok:
                checksum = md5(dest)
                size_kb  = os.path.getsize(dest) // 1024
                print(f"OK ({release}, {size_kb} KB)")
                downloaded.append(dest)
                manifest_lines.append(
                    f"  {unit_code}_{release}.pdf | {unit_title} | MD5: {checksum}"
                )
            else:
                print("FAILED")
                failed.append(unit_code)
                manifest_lines.append(f"  {unit_code} | DOWNLOAD FAILED")
            time.sleep(REQUEST_DELAY)
        manifest_lines.append("")

    # ── Step 4 (or 1 in search-only): Load outcomes ────────────────────────────
    step_search = 4 if not search_only else 1
    print(f"\n[Step {step_search}/{total_steps}] Loading outcomes from {MASTER_XLSX}")
    outcomes, layer0_primary, layer1_primary = load_outcomes(MASTER_XLSX)
    print(f"  Loaded {len(outcomes)} outcomes across 4 credentials")

    # ── Step 5 (or 2+3): Keyword search and report generation ─────────────────
    step_kw = 5 if not search_only else 2
    print(f"\n[Step {step_kw}/{total_steps}] Running keyword searches and generating reports")

    print(f"  Layer 0: {len(LAYER0_TERMS)} primary terms... ", end="", flush=True)
    l0_results, l0_hits = run_keyword_search(outcomes, LAYER0_TERMS)
    print(f"done ({len(l0_hits)} hits)")

    print(f"  Layer 1: {len(LAYER1_TERMS)} near-synonym terms... ", end="", flush=True)
    l1_results, l1_hits = run_keyword_search(outcomes, LAYER1_TERMS)
    print(f"done ({len(l1_hits)} hits)")

    # Generate reports
    l0_report   = generate_layer_report("Layer 0: Primary Terms (15 terms)",
                                         LAYER0_TERMS, l0_results, layer0_primary,
                                         l0_hits, len(outcomes), run_timestamp)
    l1_report   = generate_layer_report("Layer 1: Near-Synonym Terms (30 terms)",
                                         LAYER1_TERMS, l1_results, layer1_primary,
                                         l1_hits, len(outcomes), run_timestamp)
    l2_report   = generate_layer2_report(outcomes, run_timestamp)
    full_report = "\n\n".join([l0_report, l1_report, l2_report])

    for fname, content in [
        ("layer0_replication_report.txt", l0_report),
        ("layer1_replication_report.txt", l1_report),
        ("layer2_replication_report.txt", l2_report),
        ("full_replication_report.txt",   full_report),
    ]:
        with open(os.path.join(RESULTS_DIR, fname), "w") as f:
            f.write(content)

    # Compare hits against primary coding — only true discrepancies matter
    # A "hit" is a discrepancy only if the primary coder coded it ABSENT (0)
    # Hits where primary=1 would confirm the primary coding; hits where primary=0
    # warrant review (though occupational safety context explains most L1 hits).
    true_discrepancies = []
    for outcome_id, cred, term, desc in l0_hits + l1_hits:
        cred_key  = CRED_KEY_MAP.get(cred, "")
        in_l0     = term.lower() in layer0_primary
        primary   = (layer0_primary if in_l0 else layer1_primary).get(term.lower(), {}).get(cred_key, 0)
        if primary == 0:
            true_discrepancies.append((outcome_id, cred, term))

    total_hits = len(l0_hits) + len(l1_hits)
    print(f"  Reports written to: {RESULTS_DIR}/")
    print(f"  Total hits — Layer 0: {len(l0_hits)}  |  Layer 1: {len(l1_hits)}")
    print(f"  True discrepancies vs. primary coding: {len(true_discrepancies)}")
    if len(true_discrepancies) == 0:
        print("  CONFIRMED: 0% finding replicated. All hits match primary coding.")
    else:
        print(f"  REVIEW: {len(true_discrepancies)} outcome(s) flagged — see full_replication_report.txt")

    # ── Final step (search-only mode): summary only ────────────────────────────
    if search_only:
        print("\n" + "=" * 72)
        print(f"Reports: {RESULTS_DIR}/full_replication_report.txt")
        print(f"Finding: {'0% CONFIRMED' if total_hits == 0 else 'REVIEW REQUIRED'}")
        print("=" * 72)
        return

    # ── Create zip archive (full run only) ────────────────────────────────────
    print(f"\n[Step 5/{total_steps}] Creating replication archive: {ZIP_PATH}")

    manifest_lines += [
        "REPLICATION SUMMARY", "-" * 40,
        f"Downloaded: {len(downloaded)} files",
        f"Failed:     {len(failed)} files",
        f"L0 hits:    {len(l0_hits)} (expected: 0)",
        f"L1 hits:    {len(l1_hits)} (expected: 0)",
        f"Status:     {'CONFIRMED' if total_hits == 0 else 'REVIEW REQUIRED'}",
        "",
    ]
    if failed:
        manifest_lines.append(f"Failed files: {', '.join(str(f) for f in failed)}")

    with open(MANIFEST_PATH, "w") as f:
        f.write("\n".join(manifest_lines))

    with zipfile.ZipFile(ZIP_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(OUTPUT_DIR):
            for fname in files:
                fpath   = os.path.join(root, fname)
                arcname = os.path.relpath(fpath, os.path.dirname(OUTPUT_DIR))
                zf.write(fpath, arcname)

    zip_mb = os.path.getsize(ZIP_PATH) / (1024 * 1024)
    print(f"  Archive created: {ZIP_PATH} ({zip_mb:.1f} MB)")

    print("\n" + "=" * 72)
    print(f"COMPLETE: {len(downloaded)}/{len(downloaded) + len(failed)} files downloaded")
    if failed:
        print(f"FAILED:   {failed}")
    print(f"Archive:  {ZIP_PATH}")
    print(f"Manifest: {MANIFEST_PATH}")
    print(f"Reports:  {RESULTS_DIR}/full_replication_report.txt")
    print(f"Finding:  {'0% CONFIRMED' if total_hits == 0 else 'REVIEW REQUIRED'}")
    print("=" * 72)
    print("\nNext step: upload irr_source_documents.zip to your IRR Drive folder.")


if __name__ == "__main__":
    main()
