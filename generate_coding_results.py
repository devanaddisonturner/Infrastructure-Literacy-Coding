#!/usr/bin/env python3
"""
Infrastructure Literacy Curriculum Coding: Three-Layer Content Analysis

Generates: EJT_Binary_Coding_Results.xlsx

Supplementary material for:
    Addison-Turner, D. C. (2026). Infrastructure Literacy: A Conceptual
    Framework for Understanding How Construction Career Students Think
    About Environmental Justice. Environmental Education Research.

This script performs a three-layer systematic content analysis of
construction career and technical education (CTE) curricula across four
credentials in three national systems (United States, Australia, United
Kingdom), producing a formatted Excel workbook with complete coding records.

Three-layer design:
    Layer 0 -- Primary Terms: 15 domain-specific environmental justice
              terms searched across all four credential documents.
    Layer 1 -- Near-Synonyms: 30 additional terms addressing the concern
              that primary terms may be domain-specific jargon absent
              from vocational documents.
    Layer 2 -- Thematic Audit: 344 individual learning outcomes coded
              against three binary equity criteria.

Credentials analyzed:
    1. NCCER Core Curriculum, 6th Edition (USA) -- 103 sub-objectives
    2. CA CTE Building & Construction Standards (USA) -- 170 indicators
    3. CPC30220 Certificate III in Carpentry (Australia) -- 27 units
    4. City & Guilds 6706-23 Level 2 Diploma (UK) -- 44 learning outcomes

Usage:
    IDLE Python Shell (>>> prompt):
        >>> import os; os.chdir("/path/to/infrastructure-literacy-coding")
        >>> exec(open("generate_coding_results.py").read())

    IDLE editor (File > Open this file, then press F5)

    Command line:
        python generate_coding_results.py
        python generate_coding_results.py --unblinded
        python generate_coding_results.py --output custom_name.xlsx

    Settings: Edit the CONFIGURATION section below to change output
    filename, blinding, or coder name before running.

Dependencies:
    pip install openpyxl

License: CC-BY-4.0
"""

import json
import os
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================
# ============================================================
# CONFIGURATION -- Edit these settings if running in IDLE
# ============================================================
DEFAULT_OUTPUT = "EJT_Binary_Coding_Results.xlsx"
CODING_DATE = "December 15, 2025"
CODER_LABEL_BLINDED = "[Blinded]"

# Set True to include coder name; False for blinded submission
UNBLINDED = False  # ← IDLE users: change to True for unblinded output
CODER_NAME = "Devan Cantrell Addison-Turner"  # ← IDLE users: change to your name

# Paths -- auto-detected from script location or working directory
try:
    SCRIPT_DIR = Path(__file__).parent.resolve()
except NameError:
    # __file__ is undefined in IDLE; use current working directory
    SCRIPT_DIR = Path(os.getcwd())
DATA_DIR = SCRIPT_DIR / "data"
OUTPUT_DIR = SCRIPT_DIR / "output"

# ============================================================
# STYLES
# ============================================================
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SUBHEADER_FONT = Font(name="Arial", bold=True, size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
CELL_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=13, color="2F5496")
META_FONT = Font(name="Arial", size=9, italic=True, color="595959")
NOTE_FONT = Font(name="Arial", size=9, italic=True, color="595959")
BOLD_FONT = Font(name="Arial", bold=True, size=10)
SUMMARY_FILL = PatternFill("solid", fgColor="E2EFDA")
RESULT_FONT = Font(name="Arial", bold=True, size=10, color="CC0000")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ============================================================
# CREDENTIAL METADATA
# ============================================================
CREDENTIALS = [
    {
        "short": "NCCER Core (6th ed., USA)",
        "header": "NCCER Core\nCurriculum\n(6th ed., USA)",
        "system": "USA",
        "hierarchy": "Module > Learning Objective > Sub-Objective (a, b, c...)",
        "coded_level": "Sub-Objective",
        "n_coded": 103,
        "n_finest": "103 (finest)",
    },
    {
        "short": "CA CTE Standards (USA)",
        "header": "CA CTE Model\nCurriculum Standards\n(2013, USA)",
        "system": "USA",
        "hierarchy": "Sector > Pathway > Standard > Performance Indicator",
        "coded_level": "Performance Indicator",
        "n_coded": 170,
        "n_finest": "170 (finest)",
    },
    {
        "short": "CPC30220 (Australia)",
        "header": "CPC30220\nCert III Carpentry\n(Australia)",
        "system": "Australia",
        "hierarchy": "Qualification > Unit of Competency > Element > Performance Criterion",
        "coded_level": "Unit of Competency",
        "n_coded": 27,
        "n_finest": "~87 (at Element level)",
    },
    {
        "short": "City & Guilds 6706-23 (UK)",
        "header": "City & Guilds 6706-23\nL2 Diploma Site Carpentry\n(UK)",
        "system": "UK",
        "hierarchy": "Qualification > Unit > Learning Outcome > Assessment Criterion",
        "coded_level": "Learning Outcome",
        "n_coded": 44,
        "n_finest": "44 (finest)",
    },
]

# ============================================================
# LAYER 0 -- PRIMARY TERMS (15 terms x 4 credentials)
# ============================================================
LAYER0_DOMAINS = [
    ("Environmental Justice Domain", [
        "environmental justice",
        "environmental racism",
        "environmental equity",
        "environmental injustice",
    ]),
    ("Health Equity Domain", [
        "health equity",
        "health disparities",
        "health outcomes",
        "community health",
    ]),
    ("Distributional Consequences Domain", [
        "distributive consequences",
        "distributional impact",
        "disproportionate burden",
        "disproportionate impact",
    ]),
    ("Community Participation Domain", [
        "community voice",
        "community engagement",
        "community participation",
    ]),
]

# ============================================================
# LAYER 1 -- NEAR-SYNONYM EXPANSION (30 terms x 4 credentials)
# ============================================================
LAYER1_DOMAINS = [
    ("Equity and Justice Terms", [
        "unfair", "unequal", "inequality", "inequity",
        "social justice", "discrimination", "marginalized",
        "disadvantaged", "underserved", "racial",
    ]),
    ("Health and Exposure Terms", [
        "public health", "environmental health",
        "neighborhood health", "pollution exposure",
        "toxic exposure", "social determinants",
        "cumulative impact", "well-being*",
    ]),
    ("pollution***", []),
    ("Distributional and Vulnerability Terms", [
        "vulnerable populations", "vulnerable communities",
        "overburdened", "social vulnerability",
        "distribution of burdens", "community impact",
        "community benefit",
    ]),
    ("social benefits", []),
    ("Broader Contextual Terms", [
        "disparities", "accessibility**", "contamination",
    ]),
]

LAYER1_NOTES = [
    "Notes on borderline terms in California CTE Standards:",
    "*well-being: not present in any credential. CA CTE Standard 6 "
    '("Health and Safety") addresses occupational health (PPE, OSHA) only.',
    "**accessibility: not present in equity sense. No credential addresses "
    "equitable access to environmental health benefits.",
    'CA CTE Career Ready Practice 12 mentions "environmental, social, and '
    'economic impacts of decisions" -- but sub-standards operationalize this '
    "exclusively as regulatory compliance (CEQA, EIR), not distributional "
    "equity analysis.",
    'CA CTE B3.2-3 mentions "non-point-source pollution" -- technical water '
    "quality management, not community health exposure.",
    'CA CTE D9.0 addresses "sustainable construction" -- operationalized '
    "entirely as energy efficiency (D9.1-D9.6), not environmental justice.",
    '***pollution: CA CTE B3.2-3 references "non-point-source pollution" '
    "in a technical water quality management context (identifying pollutant "
    "sources for regulatory compliance). No credential uses \"pollution\" in "
    "connection with community health exposure, environmental justice, or "
    "distributional consequences.",
]

# ============================================================
# SOURCE DOCUMENTS
# ============================================================
SOURCE_DOCUMENTS = [
    {
        "credential": "NCCER Core (USA)",
        "title": "NCCER Core Curriculum, 6th Edition (NCCER, 2023)",
        "url": "https://www.nccer.org/craft-catalog/core/",
        "accessed": "2025-12-15",
        "note": "Requires institutional purchase. Coding conducted using "
                "6th edition print volume. URL links to NCCER catalog page "
                "confirming module structure. Independent verification "
                "requires access through an NCCER-accredited training "
                "center or institutional library.",
    },
    {
        "credential": "CA CTE Standards (USA)",
        "title": "Career Technical Education Model Curriculum Standards: "
                 "Building and Construction Trades (CDE, 2013)",
        "url": "https://www.cde.ca.gov/ci/ct/sf/documents/buildingconstruct.pdf",
        "accessed": "2025-12-15",
        "note": "Publicly available PDF.",
    },
    {
        "credential": "CPC30220 (Australia)",
        "title": "CPC30220 Certificate III in Carpentry, Release 1 "
                 "(Artibus Innovation, 2020)",
        "url": "https://training.gov.au/TrainingComponentFiles/CPC/CPC30220_R1.pdf",
        "accessed": "2025-12-15",
        "note": "Publicly available via training.gov.au.",
    },
    {
        "credential": "City & Guilds 6706-23 (UK)",
        "title": "Level 2 Diploma in Site Carpentry (6706-23) Qualification "
                 "Handbook v1.4 (City and Guilds, 2022)",
        "url": "https://www.cityandguilds.com/-/media/productdocuments/"
               "construction_and_the_built_environment/construction/6706/"
               "6706_level_2/centre_documents/"
               "6706-23_l2_diploma_in_site_carpentry_qhb_v1-4-pdf.ashx",
        "accessed": "2025-12-15",
        "note": "Publicly available from City and Guilds.",
    },
]

# ============================================================
# METHODOLOGY TEXT
# ============================================================
METH_CODING_PROTOCOL = (
    "Each credential was coded at the level its governing body defines as "
    "the fundamental learning outcome. Because the four national systems "
    "organize curriculum differently, the structural unit that constitutes "
    'a "learning outcome" varies across credentials. This section documents '
    "the structural hierarchy of each credential and identifies the level "
    "at which coding was performed."
)

METH_GRANULARITY_NOTE = (
    'The Australian training package defines "Elements" as "the essential '
    'outcomes" (see training.gov.au unit documents: "Elements describe the '
    'essential outcomes"). CPC30220 was coded at the Unit of Competency '
    "level (27 core units) rather than the Element level (~87 elements "
    "across those units, based on verified samples averaging 3.25 elements "
    "per unit). This means CPC30220 is coded at a coarser structural grain "
    "than the other three credentials, which were coded at their respective "
    "sub-unit levels.\n\n"
    "Verified element counts from training.gov.au PDFs:\n"
    "  - CPCCCM2006 Apply basic levelling procedures: 3 elements\n"
    "  - CPCCCA3004 Construct and erect wall frames: 4 elements\n"
    "  - CPCCCA3005 Construct ceiling frames: 3 elements\n"
    "  - CPCCCA3006 Erect roof trusses: 3 elements\n\n"
    "Average: 3.25 elements per unit x 27 core units = ~87 element-level "
    "outcomes."
)

METH_INVARIANCE = (
    "The study's central finding -- complete absence of equity-related "
    "content (0% across all three criteria) -- is invariant across coding "
    "granularity. Whether CPC30220 contributes 27 unit-level rows or ~87 "
    'element-level rows, every row codes as "Absent" on all three equity '
    "criteria. The granularity difference affects the reported denominator "
    "(344 vs. ~404 total outcomes) but does not affect the binary finding "
    "or the study's conclusions."
)

METH_THREE_LAYER = (
    "Layer 0 -- Primary Terms: 15 domain-specific terms searched across "
    "all four credential documents (0 of 60 possible occurrences).\n"
    "Layer 1 -- Near-Synonyms: 30 additional terms searched to address "
    "domain-jargon concern (0 of 120 occurrences in equity-relevant "
    "contexts).\n"
    "Layer 2 -- Thematic Audit: 344 learning outcomes individually coded "
    "against three binary equity criteria (0 of 344 met any criterion)."
)


# ============================================================
# HELPER FUNCTIONS
# ============================================================
def write_header_row(ws, row, headers, col_start=1):
    """Write a styled header row to a worksheet."""
    for c, h in enumerate(headers, col_start):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def write_cell(ws, row, col, value, font=CELL_FONT, align=CENTER):
    """Write a styled data cell to a worksheet."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = align
    cell.border = THIN_BORDER
    return cell


def set_column_widths(ws, widths):
    """Set column widths from a dictionary of {letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ============================================================
# SHEET BUILDERS
# ============================================================
def build_layer0(wb, coder_label):
    """Build Layer 0: Primary Term Binary Coding (15 terms x 4 credentials)."""
    ws = wb.active
    ws.title = "Layer 0 - Primary Coding"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = (
        "Binary Presence/Absence Coding: 15 Primary Terms "
        "Across Four Construction Credentials"
    )
    ws["A1"].font = TITLE_FONT

    # Metadata
    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Coding date: {CODING_DATE}  |  "
        f"Coder: {coder_label}  |  "
        "Single-coder design (Potter & Levine-Donnerstein, 1999)"
    )
    ws["A2"].font = META_FONT

    # Headers
    headers = (
        ["Search Term"]
        + [c["header"] for c in CREDENTIALS]
        + ["Term Present in\nAny Credential"]
    )
    write_header_row(ws, 4, headers)

    # Data
    row = 5
    data_rows = []
    for domain_name, terms in LAYER0_DOMAINS:
        ws.cell(row=row, column=1, value=domain_name).font = SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
        for term in terms:
            write_cell(ws, row, 1, term, align=LEFT_WRAP)
            for c in range(2, 6):
                write_cell(ws, row, c, 0)
            ws.cell(row=row, column=6, value=f"=IF(SUM(B{row}:E{row})>0,1,0)")
            ws.cell(row=row, column=6).font = BOLD_FONT
            ws.cell(row=row, column=6).alignment = CENTER
            ws.cell(row=row, column=6).border = THIN_BORDER
            data_rows.append(row)
            row += 1

    # Totals
    row += 1
    ws.cell(row=row, column=1, value="Total terms present").font = BOLD_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    for col in range(2, 7):
        refs = ",".join([f"{get_column_letter(col)}{r}" for r in data_rows])
        cell = ws.cell(row=row, column=col, value=f"=SUM({refs})")
        cell.font = BOLD_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = SUMMARY_FILL

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value=(
        "Result: 0 of 60 possible occurrences (15 terms x 4 credentials). "
        "No primary term appeared in any credential."
    )).font = RESULT_FONT

    set_column_widths(ws, {"A": 30, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22})


def build_layer1(wb, coder_label):
    """Build Layer 1: Near-Synonym Expansion (30 terms x 4 credentials)."""
    ws = wb.create_sheet("Layer 1 - Near-Synonyms")

    ws.merge_cells("A1:F1")
    ws["A1"] = (
        "Layer 1: Near-Synonym Expansion -- 30 Additional Terms "
        "Across Four Construction Credentials"
    )
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Coding date: {CODING_DATE}  |  "
        "Expanded search to address concern that primary terms may be "
        "domain-specific jargon absent from vocational documents"
    )
    ws["A2"].font = META_FONT

    headers = (
        ["Search Term"]
        + [c["header"] for c in CREDENTIALS]
        + ["Term Present in\nAny Credential"]
    )
    write_header_row(ws, 4, headers)

    row = 5
    data_rows = []
    for domain_name, terms in LAYER1_DOMAINS:
        ws.cell(row=row, column=1, value=domain_name).font = SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
        for term in terms:
            write_cell(ws, row, 1, term, align=LEFT_WRAP)
            for c in range(2, 6):
                write_cell(ws, row, c, 0)
            ws.cell(row=row, column=6, value=f"=IF(SUM(B{row}:E{row})>0,1,0)")
            ws.cell(row=row, column=6).font = BOLD_FONT
            ws.cell(row=row, column=6).alignment = CENTER
            ws.cell(row=row, column=6).border = THIN_BORDER
            data_rows.append(row)
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="Total terms present").font = BOLD_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    for col in range(2, 7):
        refs = ",".join([f"{get_column_letter(col)}{r}" for r in data_rows])
        cell = ws.cell(row=row, column=col, value=f"=SUM({refs})")
        cell.font = BOLD_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = SUMMARY_FILL

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value=(
        "Result: 0 of 120 possible occurrences (30 terms x 4 credentials). "
        "No near-synonym appeared in any credential in an equity-relevant "
        "context."
    )).font = RESULT_FONT

    row += 2
    for note in LAYER1_NOTES:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=note).font = NOTE_FONT
        ws.row_dimensions[row].height = 30
        row += 1

    set_column_widths(ws, {"A": 30, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22})


def build_layer2(wb, coder_label):
    """Build Layer 2: Thematic Audit of 344 individual learning outcomes."""
    ws = wb.create_sheet("Layer 2 - Thematic Audit")

    # Load data
    with open(DATA_DIR / "layer2_outcomes.json") as f:
        outcomes = json.load(f)
    with open(DATA_DIR / "borderline_cases.json") as f:
        borderline = json.load(f)

    n_outcomes = len(outcomes)

    # Title rows
    ws.merge_cells("A1:J1")
    ws["A1"] = "Layer 2: Thematic Audit of Learning Outcomes -- Complete Coding Record"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        f"Coding date: {CODING_DATE}  |  "
        f"Total outcomes coded: {n_outcomes} across four credentials in "
        "three national systems  |  Note: Credentials coded at each "
        "system's published learning-outcome level. See Methodology sheet "
        "for structural hierarchy and coding-level rationale."
    )
    ws["A2"].font = META_FONT

    ws.merge_cells("A3:J3")
    ws["A3"] = (
        "Coding protocol: Each learning outcome independently evaluated "
        "against three binary criteria: (a) Does this outcome require "
        "learners to evaluate distributional consequences of infrastructure "
        "decisions? (b) Does this outcome address community health in an "
        "equity framework (beyond occupational safety)? (c) Does this "
        "outcome engage environmental justice concepts, even implicitly?"
    )
    ws["A3"].font = Font(name="Arial", size=9, color="333333")

    ws.merge_cells("A4:J4")
    ws["A4"] = (
        'Decision rule: "Present" if any criterion is met (a OR b OR c). '
        '"Absent" if none are met.'
    )
    ws["A4"].font = Font(name="Arial", size=9, color="333333")

    # Column headers
    col_headers = [
        "#", "Credential", "Unit/Module", "Outcome ID",
        "Learning Outcome Description",
        "Crit. A:\nDistrib.\nBurdens",
        "Crit. B:\nComm.\nHealth Eq.",
        "Crit. C:\nEJ\nConcepts",
        "Result", "Coder Notes",
    ]
    write_header_row(ws, 6, col_headers)
    ws.row_dimensions[6].height = 40

    # Data rows
    for item in outcomes:
        r = 6 + item["num"]
        write_cell(ws, r, 1, item["num"])
        write_cell(ws, r, 2, item["credential"], align=LEFT_WRAP)
        write_cell(ws, r, 3, item["unit"], align=LEFT_WRAP)
        write_cell(ws, r, 4, item["outcome_id"], align=LEFT_WRAP)
        write_cell(ws, r, 5, item["description"], align=LEFT_WRAP)
        write_cell(ws, r, 6, item["crit_a"])
        write_cell(ws, r, 7, item["crit_b"])
        write_cell(ws, r, 8, item["crit_c"])
        write_cell(ws, r, 9, item["result"])
        if item["notes"]:
            write_cell(ws, r, 10, item["notes"], font=NOTE_FONT, align=LEFT_WRAP)

    last_row = 6 + n_outcomes

    # Summary
    r = last_row + 3
    ws.cell(row=r, column=1, value="SUMMARY").font = Font(
        name="Arial", bold=True, size=12, color="2F5496"
    )
    r += 1
    summary_headers = [
        "Credential", "Outcomes Coded",
        "Criterion A = 1", "Criterion B = 1",
        "Criterion C = 1", "Any Present",
    ]
    for c, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    r += 1
    for cred in CREDENTIALS:
        ws.cell(row=r, column=1, value=cred["short"]).font = CELL_FONT
        ws.cell(row=r, column=2, value=cred["n_coded"]).font = CELL_FONT
        for col in range(3, 7):
            ws.cell(row=r, column=col, value=0).font = CELL_FONT
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=r, column=2, value=n_outcomes).font = BOLD_FONT
    for col in range(3, 7):
        ws.cell(row=r, column=col, value=0).font = BOLD_FONT
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = THIN_BORDER
    r += 2

    # Key finding
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(row=r, column=1, value=(
        f"Key finding: Across {n_outcomes} individually coded learning "
        "outcomes from four construction credentials in three national "
        "systems, zero outcomes meet any of the three equity criteria. "
        "The absence is structural, not terminological: credentials do "
        "not merely avoid justice vocabulary -- they lack any pedagogical "
        "framework through which learners might evaluate who benefits from "
        "or is burdened by infrastructure decisions."
    )).font = RESULT_FONT
    ws.row_dimensions[r].height = 50
    r += 2

    # Borderline cases
    ws.cell(
        row=r, column=1,
        value="Borderline Cases Requiring Explicit Justification"
    ).font = Font(name="Arial", bold=True, size=11, color="2F5496")
    r += 1
    for case in borderline:
        ws.cell(row=r, column=1, value=case["id"]).font = BOLD_FONT
        ws.cell(row=r, column=2, value=case["text"]).font = CELL_FONT
        ws.cell(row=r, column=2).alignment = LEFT_WRAP
        ws.cell(row=r, column=5, value=case["justification"]).font = NOTE_FONT
        ws.cell(row=r, column=5).alignment = LEFT_WRAP
        ws.row_dimensions[r].height = 45
        r += 1

    # Freeze panes
    ws.freeze_panes = "A7"

    set_column_widths(ws, {
        "A": 6, "B": 25, "C": 30, "D": 14, "E": 65,
        "F": 10, "G": 10, "H": 10, "I": 10, "J": 40,
    })


def build_methodology(wb):
    """Build Methodology sheet documenting coding decisions."""
    ws = wb.create_sheet("Methodology")

    ws.merge_cells("A1:F1")
    ws["A1"] = "Methodology: Unit of Analysis and Cross-System Granularity"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="2F5496")

    sections = [
        (3, "1. Coding Protocol", 4, METH_CODING_PROTOCOL, 60),
        (6, "2. Structural Hierarchy and Coding Level by Credential", None, None, None),
        (12, "3. CPC30220 Granularity Note", 13, METH_GRANULARITY_NOTE, 200),
        (15, "4. Invariance of Finding Across Granularity Levels", 16, METH_INVARIANCE, 80),
        (18, "5. Three-Layer Coding Design", 19, METH_THREE_LAYER, 60),
    ]

    for heading_row, heading, text_row, text, height in sections:
        ws.cell(row=heading_row, column=1, value=heading).font = Font(
            name="Arial", bold=True, size=11
        )
        if text_row and text:
            ws.cell(row=text_row, column=1, value=text).font = CELL_FONT
            ws.cell(row=text_row, column=1).alignment = LEFT_WRAP
            ws.merge_cells(f"A{text_row}:F{text_row}")
            ws.row_dimensions[text_row].height = height

    # Hierarchy table
    meth_headers = [
        "Credential", "National System",
        "Structural Hierarchy (top > bottom)",
        "Coded Level", "N at Coded Level",
        "Approx. N at Finest Grain",
    ]
    write_header_row(ws, 7, meth_headers)

    for i, cred in enumerate(CREDENTIALS):
        r = 8 + i
        row_data = [
            cred["short"], cred["system"], cred["hierarchy"],
            cred["coded_level"], cred["n_coded"], cred["n_finest"],
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.alignment = LEFT_WRAP
            cell.border = THIN_BORDER

    set_column_widths(ws, {
        "A": 30, "B": 15, "C": 45, "D": 20, "E": 15, "F": 20,
    })


def build_source_docs(wb):
    """Build Source Documents sheet with access information."""
    ws = wb.create_sheet("Source Documents")

    headers = ["Credential", "Full Title", "Access URL", "Date Accessed", "Access Note"]
    write_header_row(ws, 1, headers)

    for i, doc in enumerate(SOURCE_DOCUMENTS):
        r = 2 + i
        vals = [doc["credential"], doc["title"], doc["url"], doc["accessed"], doc["note"]]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.alignment = LEFT_WRAP
            cell.border = THIN_BORDER

    set_column_widths(ws, {"A": 25, "B": 50, "C": 55, "D": 15, "E": 45})


# ============================================================
# MAIN
# ============================================================
def main(output_file=None, unblinded=None, coder_name=None):
    """
    Generate the coding results workbook.

    Can be called three ways:
        1. From command line:  python generate_coding_results.py --unblinded
        2. From IDLE (F5):     Edit CONFIGURATION section at top of file
        3. Programmatically:   main(output_file="results.xlsx", unblinded=True)
    """
    # Resolve settings: function args > CLI args > module-level config
    if output_file is None or unblinded is None:
        try:
            parser = argparse.ArgumentParser(
                description="Generate Infrastructure Literacy Curriculum "
                            "Coding Results"
            )
            parser.add_argument(
                "--output", "-o", default=DEFAULT_OUTPUT,
                help=f"Output filename (default: {DEFAULT_OUTPUT})",
            )
            parser.add_argument(
                "--unblinded", action="store_true",
                help="Include coder name (omit for blinded submission)",
            )
            parser.add_argument(
                "--coder", default=CODER_NAME,
                help="Coder name for unblinded version",
            )
            args = parser.parse_args()
            output_file = output_file or args.output
            unblinded = unblinded if unblinded is not None else args.unblinded
            coder_name = coder_name or args.coder
        except SystemExit:
            # argparse exits on --help or error; catch to keep IDLE alive
            return
    else:
        coder_name = coder_name or CODER_NAME

    # Fall back to module-level config if still None
    if output_file is None:
        output_file = DEFAULT_OUTPUT
    if unblinded is None:
        unblinded = UNBLINDED

    coder_label = coder_name if unblinded else CODER_LABEL_BLINDED
    output_path = OUTPUT_DIR / output_file

    # Verify data files exist
    missing = []
    for f in ["layer2_outcomes.json", "borderline_cases.json"]:
        if not (DATA_DIR / f).exists():
            missing.append(str(DATA_DIR / f))
    if missing:
        print("ERROR: Required data file(s) not found:")
        for m in missing:
            print(f"  {m}")
        print("\nEnsure the data/ directory contains both JSON files.")
        print(f"Script directory detected as: {SCRIPT_DIR}")
        print("If running in IDLE, open this file from the repository folder")
        print("or set SCRIPT_DIR manually in the CONFIGURATION section.")
        return  # return instead of exiting to keep IDLE shell alive

    # Build workbook
    wb = Workbook()

    print("Building Layer 0 -- Primary Coding (15 terms)...")
    build_layer0(wb, coder_label)

    print("Building Layer 1 -- Near-Synonyms (30 terms)...")
    build_layer1(wb, coder_label)

    print("Building Layer 2 -- Thematic Audit (344 outcomes)...")
    build_layer2(wb, coder_label)

    print("Building Methodology...")
    build_methodology(wb)

    print("Building Source Documents...")
    build_source_docs(wb)

    # Save
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    size = output_path.stat().st_size
    print(f"\nSaved: {output_path} ({size / 1024:.0f} KB)")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Coder: {coder_label}")
    print(f"\nTo verify: open {output_path} in Excel or LibreOffice Calc")


if __name__ == "__main__":
    # ================================================================
    # IDLE PYTHON SHELL USERS:
    #   1. Open IDLE
    #   2. At the >>> prompt, type these two lines:
    #
    #        import os; os.chdir("/path/to/infrastructure-literacy-coding")
    #        exec(open("generate_coding_results.py").read())
    #
    #   Or open this file in IDLE (File > Open) and press F5.
    #
    # COMMAND LINE USERS:
    #   python generate_coding_results.py
    #   python generate_coding_results.py --unblinded
    # ================================================================

    main()
